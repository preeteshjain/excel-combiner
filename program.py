import os
import pandas

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None, truncate_sheet=False, **to_excel_kwargs):
    from openpyxl import load_workbook

    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pandas.ExcelWriter(filename, engine='openpyxl')

    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        writer.book = load_workbook(filename)

        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        if truncate_sheet and sheet_name in writer.book.sheetnames:
            idx = writer.book.sheetnames.index(sheet_name)
            writer.book.remove(writer.book.worksheets[idx])
            writer.book.create_sheet(sheet_name, idx)

        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        pass

    if startrow is None:
        startrow = 0

    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
    writer.save()

directory = 'files'
main_df = pandas.DataFrame()

for filename in os.listdir(directory):
    if filename.endswith('.xlsx'):
        xls = pandas.ExcelFile('files/' + filename)
        print('Loading ' + filename + '...')
        for sheet_name in xls.sheet_names:
            data = pandas.read_excel('files/' + filename, sheet_name = sheet_name)
            append_df_to_excel('master-file.xlsx', data)
            print(filename + ' - ' + sheet_name + ' processed.')
        print('***************************')

print('--------------------------------')
print('master-file.xlsx has been generated!')
print('--------------------------------')
